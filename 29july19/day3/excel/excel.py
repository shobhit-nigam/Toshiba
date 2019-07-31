
# coding: utf-8

# In[1]:

from openpyxl import load_workbook
wba = load_workbook('random.xlsx')


# In[2]:

print(wba.sheetnames)


# In[3]:

sheeta = wba.get_sheet_by_name('tue')


# In[4]:

type(sheeta)


# In[5]:

sheeta.title


# In[7]:

sheeta['B4'].value


# In[8]:

for i in range(1, 6):
    print(i, sheeta.cell(row=i, column=3).value)


# In[9]:

import pandas as pd


# In[10]:

x1 = pd.ExcelFile('random.xlsx')


# In[11]:

x1.sheet_names


# In[12]:

dfa = x1.parse('tue')


# In[13]:

dfa


# In[14]:

dfa.shape


# In[15]:

dicta = dict(zip(dfa.title, dfa.val2))


# In[16]:

dicta


# In[17]:

dictb = {'Data':[4,5,6,7,8], 'Value':[6,7,8,9,1]}


# In[18]:

dfb = pd.DataFrame(dictb)


# In[19]:

writer = pd.ExcelWriter('foodpanda.xlsx' , engine='xlsxwriter')


# In[20]:

dfb.to_excel(writer, sheet_name = 'biryani')


# In[21]:

writer.save()


# In[23]:

type(dfa.title)


# In[25]:

from openpyxl import Workbook
wba = Workbook()
wsa = wba.active
wsa['A2'] = 22
wsa['B2'] = 23
wsa['C2'] = "=SUM(A2:B2)"
wba.save('zomato.xlsx')


# In[27]:

from openpyxl import Workbook
from openpyxl.drawing.image import Image
wba = Workbook()
wsa = wba.active
img_obj = Image('toshiba.png')
wsa.add_image(img_obj, 'F6')
wba.save('swiggy.xlsx')


# In[29]:

from openpyxl import Workbook
wba = Workbook()
wsa = wba.active
for i in range(0,20, 2):
    wsa.append([i])

from openpyxl.chart import BarChart, Reference
values = Reference(wsa, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
wsa.add_chart(chart, 'D4')
wba.save('ubereats.xlsx')


# In[ ]:



