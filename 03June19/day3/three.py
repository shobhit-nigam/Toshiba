
# coding: utf-8

# In[1]:

import openpyxl 


# In[2]:

from openpyxl import load_workbook 


# In[4]:

wba = load_workbook("sample.xlsx")


# In[5]:

print(wba.sheetnames)


# In[10]:

whole_data = wba.get_sheet_by_name('Sheet1')


# In[8]:

type(sheet)


# In[11]:

whole_data.title


# In[12]:

varx = whole_data['B3']


# In[13]:

varx


# In[14]:

print(varx)


# In[15]:

varx.value


# In[16]:

type(varx.value)


# In[17]:

varx.row


# In[18]:

whole_data['A1'].value


# In[20]:

whole_data.cell(row = 1, column = 2).value


# In[21]:

for i in range(1,5):
    print(i, whole_data.cell(row = i, column = 2).value)


# In[22]:

import pandas as pd


# In[23]:

file = "sample.xlsx"


# In[24]:

x1 = pd.ExcelFile(file)


# In[25]:

type(x1)


# In[26]:

print(x1.sheet_names)


# In[28]:




# In[29]:

df1 = x1.parse('Sheet1')


# In[30]:

type(df1)


# In[31]:

df1


# In[32]:

df1.shape


# In[33]:

key_a = ['a', 'b', 'c']


# In[34]:

val_a = [3, 7, 9]


# In[35]:

dict_a = dict(zip(key_a, val_a))


# In[36]:

print(dict_a)


# In[37]:

excel_dict1 = dict(zip(df1.Name, df1.Number_2))


# In[38]:

print(excel_dict1)


# In[39]:

df2 = pd.DataFrame({'Data' : [10,20,33,4,6,89]})


# In[41]:

write = pd.ExcelWriter('sample_b.xlsx', engine = 'xlsxwriter')


# In[42]:

df2.to_excel(write, sheet_name='mysheet')


# In[43]:

write.save()


# In[44]:

from openpyxl import Workbook


# In[45]:

from openpyxl.utils import get_column_letter


# In[46]:

wbb = Workbook()


# In[47]:

file_b = "new_book.xlsx"


# In[48]:

ws1 = wbb.active


# In[49]:

ws1.title = "chowkidar"


# In[50]:

for row in range(1, 20):
    ws1.append(range(200))


# In[59]:

#ws2 = wbb.active
ws4 = wbb.create_sheet("Pi")


# In[58]:

wbb.save(file_b)


# In[60]:

ws8 = wbb.create_sheet("new_name")


# In[61]:

ws8['B2'] = "hey"


# In[63]:

wbb.save(file_b)


# In[64]:

import datetime


# In[66]:

wbc = Workbook()


# In[67]:

wc1 = wbc.active


# In[68]:

wc1['B2'] = datetime.datetime(2019, 7, 6)


# In[69]:

wc1['B2'].number_format


# In[70]:

file_b = "test1.xlsx"


# In[71]:

wbc.save(file_b)


# In[72]:

#wc1['A2'] = "=SUM(1,1)" 
# wc1.merge_cells('A2:D2')
# wc1.unmerge_cells()
wc1['C2'] = "attaching pic"


# In[82]:

image = Image('pdf.png')


# In[80]:

from openpyxl.drawing.image import Image


# In[83]:

wc1.add_image(image, 'F4')


# In[84]:

wbc.save(file_b)


# In[85]:

wbd = Workbook()


# In[86]:

wd1 = wbd.active


# In[87]:

for i in range(8):
    wd1.append([i])


# In[88]:

from openpyxl.chart import Barchart, Series, Reference, Unit


# In[89]:

from openpyxl.chart import BarChart, Series, Reference


# In[90]:

val = Reference(wd1, min_col = 1, min_row = 1, max_col = 1, max_row = 8)


# In[91]:

chart = BarChart()


# In[92]:

chart.add_data(val)


# In[93]:

wd1.add_chart(chart, 'D10')


# In[94]:

wbd.save("graph.xlsx")


# In[95]:

import sqlite3


# In[96]:

conn = sqlite3.connect("cricket.db")


# In[97]:

qstring = "SELECT name FROM players WHERE price > 600000"


# In[98]:

df4 = pd.read_sql_query(qstring, conn)


# In[100]:

df4.shape


# In[101]:

for name in df4['name']:
    print(name)


# In[102]:

df4


# In[103]:

qstringb = "SELECT * FROM players WHERE price > 600000"


# In[104]:

df5 = pd.read_sql_query(qstringb, conn)


# In[105]:

df5


# In[ ]:



